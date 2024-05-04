import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import tkinter.messagebox as messagebox
import openpyxl

def import_excel():
    # Mở hộp thoại để chọn file Excel
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])

    # Kiểm tra xem người dùng đã chọn file chưa
    if filepath:
        # Đọc dữ liệu từ file Excel
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            # Kiểm tra xem hàng đó có chứa dữ liệu không
            if any(cell.strip() for cell in row if cell is not None):
                data.append(row)

        # Kiểm tra xem có dữ liệu từ file Excel không
        if data:
            # Sắp xếp dữ liệu theo cột "KVH index" từ lớn đến bé
            data_sorted = sorted(data[1:], key=lambda x: x[0], reverse=True)
            data_sorted.insert(0, data[0])  # Thêm header vào lại dữ liệu

            # Xóa dữ liệu cũ trong Treeview
            for i in treeview.get_children():
                treeview.delete(i)

            # Thêm các cột vào Treeview
            treeview["columns"] = data_sorted[0]
            
            # Đặt tiêu đề cho từng cột và ngăn chúng được mở rộng
            for column in treeview["columns"]:
                treeview.heading(column, text=column, anchor="center")
                treeview.column(column, stretch=False, width=100, minwidth=100)

            treeview.tag_configure("oddrow", background="#f0f0f0")
            treeview.tag_configure("evenrow", background="#ffffff")
            
            # Hiển thị dữ liệu từ file Excel lên Treeview
            global data_dict
            data_dict = []
            for i in range(len(data_sorted)):
                if i != 0:
                    treeview.insert("", "end", values=data_sorted[i])
                    data_dict.append({data_sorted[0][j]: data_sorted[i][j] for j in range(len(data_sorted[0]))})
            
            # Tô màu cho các hàng dựa trên chỉ số KVH Index và Long Hổ
            color_rows_by_kvh_index()

def export_excel():
    # Kiểm tra xem có dữ liệu để xuất không
    if 'data_dict' not in globals():
        messagebox.showwarning("Cảnh báo", "Không có dữ liệu để xuất!")
        return

    # Mở hộp thoại để chọn vị trí lưu file Excel
    filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

    # Kiểm tra xem người dùng đã chọn vị trí lưu hay chưa
    if filepath:
        # Tạo một workbook mới
        workbook = openpyxl.Workbook()
        # Chọn worksheet đầu tiên
        worksheet = workbook.active

        # Thêm tiêu đề cho các cột
        headers = list(data_dict[0].keys())
        worksheet.append(headers)

        # Thêm dữ liệu từ data_dict vào worksheet
        for row_data in data_dict:
            row_values = list(row_data.values())
            worksheet.append(row_values)

        # Lưu workbook vào file Excel
        workbook.save(filepath)
        messagebox.showinfo("Thông báo", "Xuất file Excel thành công.")

def open_statistics_window():

    # Tính tổng số KVH
    total_kvh = 0
    pcc_max = int(treeview.item(treeview.get_children()[0], "values")[27])
    pcc_min = int(treeview.item(treeview.get_children()[0], "values")[27])
    for row in treeview.get_children():
        values = treeview.item(row, "values")
        long_ho = int(values[26])
        if long_ho in [0, 1]:
            kvh = int(values[27])
            # Tìm KVH max
            if kvh > pcc_max:
                pcc_max = kvh
            # Tìm KVH min
            if kvh < pcc_min:
                pcc_min = kvh
            total_kvh += kvh
    avg_kvh = total_kvh/24

    long_total_kvh = 0
    long_max = 0
    long_min = 999
    for row in treeview.get_children():
        values = treeview.item(row, "values")
        long_ho = int(values[26])
        if long_ho in [0]:
            kvh = int(values[27])
            # Tìm long max
            if kvh > long_max:
                long_max = kvh
            # Tìm long min
            if kvh < long_min:
                long_min = kvh
            long_total_kvh += kvh
    long_avg_kvh = long_total_kvh/12

    ho_total_kvh = 0
    ho_max = 0
    ho_min = 999
    for row in treeview.get_children():
        values = treeview.item(row, "values")
        long_ho = int(values[26])
        if long_ho in [1]:
            kvh = int(values[27])
            # Tìm ho max
            if kvh > ho_max:
                ho_max = kvh
            # Tìm ho min
            if kvh < ho_min:
                ho_min = kvh
            ho_total_kvh += kvh
    ho_avg_kvh = ho_total_kvh/12
    
    # Tạo một cửa sổ con
    stats_window = tk.Toplevel(root)
    stats_window.title("Thông kê")
    
    # Khóa kích thước của cửa sổ con
    stats_window.resizable(False, False)
    
    # Tạo frame cho PCC FC
    pcc_frame = ttk.Frame(stats_window, borderwidth=2, relief="groove", padding=(10, 5))
    pcc_frame.grid(row=0, column=0, padx=10, pady=10)

    # Nhãn cho PCC FC
    ttk.Label(pcc_frame, text="PCC FC", font=("Helvetica", 14, "bold")).grid(row=0, column=0, columnspan=2, pady=(0, 10))

    # Các trường thông tin cho PCC FC
    ttk.Label(pcc_frame, text="Tổng KVH:").grid(row=1, column=0, sticky="e")
    pcc_total_kvh_label = ttk.Label(pcc_frame, text=str(total_kvh))
    pcc_total_kvh_label.grid(row=1, column=1, sticky="w")
    
    ttk.Label(pcc_frame, text="KVH Trung bình:").grid(row=2, column=0, sticky="e")
    pcc_avg_kvh_label = ttk.Label(pcc_frame, text=str(avg_kvh))
    pcc_avg_kvh_label.grid(row=2, column=1, sticky="w")

    ttk.Label(pcc_frame, text="KVH Cao nhất:").grid(row=3, column=0, sticky="e")
    pcc_max_kvh_label = ttk.Label(pcc_frame, text=str(pcc_max))
    pcc_max_kvh_label.grid(row=3, column=1, sticky="w")

    ttk.Label(pcc_frame, text="KVH Thấp nhất:").grid(row=4, column=0, sticky="e")
    pcc_min_kvh_label = ttk.Label(pcc_frame, text=str(pcc_min))
    pcc_min_kvh_label.grid(row=4, column=1, sticky="w")

    # Tạo frame cho Long FC
    long_frame = ttk.Frame(stats_window, borderwidth=2, relief="groove", padding=(10, 5))
    long_frame.grid(row=0, column=1, padx=10, pady=10)

    # Nhãn cho Long FC
    ttk.Label(long_frame, text="Long FC", font=("Helvetica", 14, "bold")).grid(row=0, column=0, columnspan=2, pady=(0, 10))

    # Các trường thông tin cho Long FC
    ttk.Label(long_frame, text="Tổng KVH:").grid(row=1, column=0, sticky="e")
    long_total_kvh_label = ttk.Label(long_frame, text=str(long_total_kvh))
    long_total_kvh_label.grid(row=1, column=1, sticky="e")

    ttk.Label(long_frame, text="KVH Trung bình:").grid(row=2, column=0, sticky="e")
    long_avg_kvh_label = ttk.Label(long_frame, text=str(long_avg_kvh))
    long_avg_kvh_label.grid(row=2, column=1, sticky="e")

    ttk.Label(long_frame, text="KVH Cao nhất:").grid(row=3, column=0, sticky="e")
    long_max_label = ttk.Label(long_frame, text=str(long_max))
    long_max_label.grid(row=3, column=1, sticky="e")

    ttk.Label(long_frame, text="KVH Thấp nhất:").grid(row=4, column=0, sticky="e")
    long_min_label = ttk.Label(long_frame, text=str(long_min))
    long_min_label.grid(row=4, column=1, sticky="e")

    # Tạo frame cho ho FC
    ho_frame = ttk.Frame(stats_window, borderwidth=2, relief="groove", padding=(10, 5))
    ho_frame.grid(row=0, column=2, padx=10, pady=10)

    # Nhãn cho ho FC
    ttk.Label(ho_frame, text="Hổ FC", font=("Helvetica", 14, "bold")).grid(row=0, column=0, columnspan=2, pady=(0, 10))

    # Các trường thông tin cho ho FC
    ttk.Label(ho_frame, text="Tổng KVH:").grid(row=1, column=0, sticky="e")
    ho_total_kvh_label = ttk.Label(ho_frame, text=str(ho_total_kvh))
    ho_total_kvh_label.grid(row=1, column=1, sticky="e")

    ttk.Label(ho_frame, text="KVH Trung bình:").grid(row=2, column=0, sticky="e")
    ho_avg_kvh_label = ttk.Label(ho_frame, text=str(ho_avg_kvh))
    ho_avg_kvh_label.grid(row=2, column=1, sticky="e")

    ttk.Label(ho_frame, text="KVH Cao nhất:").grid(row=3, column=0, sticky="e")
    ho_max_label = ttk.Label(ho_frame, text=str(ho_max))
    ho_max_label.grid(row=3, column=1, sticky="e")

    ttk.Label(ho_frame, text="KVH Thấp nhất:").grid(row=4, column=0, sticky="e")
    ho_min_label = ttk.Label(ho_frame, text=str(ho_min))
    ho_min_label.grid(row=4, column=1, sticky="e")

def sort_data_by_kvh_index():
    # Kiểm tra xem có dữ liệu để sắp xếp không
    if 'data_dict' not in globals():
        messagebox.showwarning("Cảnh báo", "Không có dữ liệu để sắp xếp!")
        return

    # Sắp xếp dữ liệu theo chỉ số "KVH index" từ lớn đến bé
    data_dict_sorted = sorted(data_dict, key=lambda x: x.get('KVH Index', 0), reverse=True)

    # Xóa dữ liệu cũ trong Treeview
    for i in treeview.get_children():
        treeview.delete(i)

    # Hiển thị dữ liệu đã sắp xếp lên Treeview
    for row_data in data_dict_sorted:
        treeview.insert("", "end", values=list(row_data.values()))

    # Tô màu cho các hàng dựa trên chỉ số KVH Index và Long Hổ
    color_rows_by_kvh_index()

def color_rows_by_kvh_index():
    for row_id in treeview.get_children():
        row_data = treeview.item(row_id)['values']
        kvh_index = int(row_data[27])  # Giả sử chỉ số KVH Index ở cột thứ 28 (đếm từ 0)
        long_ho = int(row_data[26])  # Giả sử chỉ số "Long Hổ" ở cột thứ 27 (đếm từ 0)

        # Xóa các tags cũ trước khi thêm tags mới
        treeview.item(row_id, tags=())

        # Thêm tags mới dựa trên giá trị của KVH Index và Long Hổ
        if long_ho == 2:
            treeview.item(row_id, tags=("purple", "white_text"))
        elif kvh_index > 85:
            treeview.item(row_id, tags=("green",))
        elif 80 <= kvh_index <= 85:
            treeview.item(row_id, tags=("lightblue",))
        elif 75 <= kvh_index <= 79:
            treeview.item(row_id, tags=("yellow",))
        elif 70 <= kvh_index <= 74:
            treeview.item(row_id, tags=("pink",))
        else:
            treeview.item(row_id, tags=("red",))


def option():
    if 'data_dict' not in globals():
        messagebox.showwarning("Cảnh báo", "Vui lòng nhập dữ liệu trước!")
        return
    
    # Tạo một cửa sổ con
    option_window = tk.Toplevel(root)
    option_window.title("Tùy chọn")
    option_window.geometry("700x675")

    # Khóa kích thước của cửa sổ con
    option_window.resizable(False, False)

    # Tạo một frame chứa các thông tin và căn bên phải
    main_frame = ttk.Frame(option_window)
    main_frame.pack(padx=10, pady=10, fill="both", expand=True)

    # Tạo một frame để chứa thông tin và thiết lập màu nền là trắng
    info_frame = ttk.Frame(main_frame, style="Info.TFrame")
    info_frame.pack(side="left", padx=10, pady=10, fill="both", expand=True)  # Thêm fill và expand vào đây

    # Lấy danh sách tên từ cột "Ho va ten" trong bảng
    def get_names():
        names = set()  # Sử dụng set để loại bỏ các tên trùng lặp
        for row in treeview.get_children():
            name = treeview.item(row, "values")[0]  # Lấy giá trị của cột "Ho va ten"
            names.add(name)
        return sorted(list(names))  # Chuyển set thành list và sắp xếp tên theo thứ tự bảng chữ cái

    # Tạo một Combobox để chọn tên
    combo_box = ttk.Combobox(main_frame, values=get_names(), width=30) # Đặt chiều rộng của Combobox là 30
    combo_box.pack(side="top", padx=10, pady=10)

    # Tạo một nhãn để hiển thị thông tin tương ứng với tên được chọn
    info_text = tk.Text(info_frame, height=10, width=50, state='disabled')  # Đặt state='disabled'
    info_text.pack(padx=10, pady=10, fill="both", expand=True)  # Thêm fill và expand vào đây

    def get_info_from_obj(obj):
        info = ""
        count = 0
        for key, value in obj.items():
            info += f"{key}: {value}\n"
            count += 1
            if count % 4 == 0:
                info += "----\n"
        
        return info
    
    def reset_values():
        travel_choice.set("Go")  # Reset giá trị của radio button "Đi"
        register_choice.set("Registered")  # Reset giá trị của radio button "Đăng ký"        
        fan_choice.set("No Fan")
        goal_value.set(0)
        self_injury_value.set(False) 
        injury_checkbox_value.set(False)
        wrong_clothes_checkbox_value.set(False)
        late_fee_checkbox_value.set(False)
        supervisor_checkbox_value.set(False)
        banana_checkbox_value.set(False)
        hattrick_checkbox_value.set(False)
        win_award_checkbox_value.set(False)
        coach_checkbox_value.set(False)
        yellow_card_value.set(0)
        red_card_value.set(0)
        
    # Liên kết sự kiện chọn từ Combobox với hàm xử lý
    def reset_values():
        travel_choice.set("Go")  # Reset giá trị của radio button "Đi"
        register_choice.set("Registered")  # Reset giá trị của radio button "Đăng ký"        
        fan_choice.set("No Fan")
        goal_value.set(0)
        self_injury_value.set(False) 
        injury_checkbox_value.set(False)
        wrong_clothes_checkbox_value.set(False)
        late_fee_checkbox_value.set(False)
        supervisor_checkbox_value.set(False)
        banana_checkbox_value.set(False)
        hattrick_checkbox_value.set(False)
        win_award_checkbox_value.set(False)
        coach_checkbox_value.set(False)
        yellow_card_value.set(0)
        red_card_value.set(0)
        
    # Lọc danh sách tên dựa trên ký tự nhập vào Combobox
    def filter_names(event):
        # Xóa danh sách tên hiện tại trong Combobox
        combo_box["values"] = ()

        # Lấy danh sách tất cả các tên
        all_names = get_names()

        # Lọc danh sách tên để chỉ hiển thị những tên chứa ký tự đã nhập vào Combobox
        filtered_names = [name for name in all_names if combo_box.get().lower() in name.lower()]

        # Cập nhật danh sách tên mới vào Combobox
        combo_box["values"] = filtered_names

    # Liên kết sự kiện nhập liệu vào Combobox với hàm xử lý lọc danh sách tên
    combo_box.bind("<KeyRelease>", filter_names)

    # Khi chọn một tên từ Combobox, hiển thị thông tin tương ứng lên Text widget
    def show_info(event):
        # Bỏ chọn các mục trong Treeview
        for row in treeview.selection():
            treeview.selection_remove(row)
        
        # Xóa nội dung hiện có trong Text widget
        info_text.config(state='normal')
        info_text.delete(1.0, "end")

        # Lấy tên được chọn từ Combobox
        selected_name = combo_box.get()

        # Lấy thông tin tương ứng từ data_dict
        for obj in data_dict:
            if obj["Họ và tên"] == selected_name:
                info_text.insert("end", get_info_from_obj(obj))
        
        # Khóa lại Text widget để ngăn người dùng chỉnh sửa
        info_text.config(state='disabled')

    # Liên kết sự kiện chọn một tên từ Combobox với hàm hiển thị thông tin
    combo_box.bind("<<ComboboxSelected>>", show_info)

    def combo_box_selected(event):
        reset_values()  # Đảm bảo rằng radio buttons được reset mỗi khi thay đổi giá trị trong combobox
        selected_name = combo_box.get()
        info_obj = get_info(selected_name)
        info = get_info_from_obj(info_obj)
        info_text.config(state='normal')  # Cho phép chỉnh sửa trạng thái
        info_text.delete(1.0, tk.END)  # Xóa nội dung cũ trước khi thêm mới
        info_text.insert(tk.END, info)
        info_text.config(state='disabled')  # Khóa lại trạng thái

    combo_box.bind("<<ComboboxSelected>>", combo_box_selected)

    # Tạo một frame để chứa radio button và căn sang bên phải
    radio_frame = ttk.Frame(main_frame)
    radio_frame.pack(side="right", fill="both", expand=True, padx=10, pady=10)  # Thêm fill và expand vào đây

    # Biến để lưu giá trị của radio button
    global travel_choice
    travel_choice = tk.StringVar()
    global register_choice 
    register_choice = tk.StringVar()
    global fan_choice
    fan_choice = tk.StringVar()
    global goal_value
    goal_value = tk.IntVar(value=0)
    global self_injury_value
    self_injury_value = tk.BooleanVar(value=False)
    global injury_checkbox_value
    injury_checkbox_value = tk.BooleanVar(value=False)  
    global wrong_clothes_checkbox_value
    wrong_clothes_checkbox_value = tk.BooleanVar(value=False)  
    global late_fee_checkbox_value
    late_fee_checkbox_value = tk.BooleanVar(value=False)
    global supervisor_checkbox_value
    supervisor_checkbox_value = tk.BooleanVar(value=False)
    global banana_checkbox_value
    banana_checkbox_value = tk.BooleanVar(value=False)
    global hattrick_checkbox_value
    hattrick_checkbox_value = tk.BooleanVar(value=False)
    global win_award_checkbox_value
    win_award_checkbox_value = tk.BooleanVar(value=False)
    global coach_checkbox_value
    coach_checkbox_value = tk.BooleanVar(value=False)
    global yellow_card_value
    yellow_card_value = tk.IntVar(value=0)
    global red_card_value
    red_card_value = tk.IntVar(value=0)
    
    # Lấy thông tin tương ứng với tên từ bảng
    def get_info(selected_name):
        for row in data_dict:
            name = row["Họ và tên"]
            long_ho = row["Long Hổ"]
            if name == selected_name:
                info_obj = {}
                count = 0
                for key, value in row.items():
                    if key.startswith("Vòng"):
                        key = "Vòng"
                        
                    info_obj[key.strip()] = value
                    count += 1 
                    
                info_obj["Bàn thắng"] += goal_value.get()
                
                if injury_checkbox_value.get():
                    info_obj["Chấn Thg"] -= 1
                
                if self_injury_value.get():
                    info_obj["Bạn CTg"] -= 2
                    
                if wrong_clothes_checkbox_value.get():
                    info_obj["Áo Ko Đúng"] -= 1
                    
                if late_fee_checkbox_value.get():
                    info_obj["Phí muộn"] -= 2
                
                if supervisor_checkbox_value.get():
                    info_obj["Giám hộ"] += 1
                
                if banana_checkbox_value.get():
                    info_obj["Chuối"] += 1
                    
                if hattrick_checkbox_value.get():
                    info_obj["Hattrick"] += 1
                
                if win_award_checkbox_value.get():
                    info_obj["Thưởng thắng"] += 1
                    info_obj["KVH Index"] += 1
                
                if coach_checkbox_value.get():
                    info_obj["HLV"] += 1
                
                
                info_obj["Yellow Card"] -= yellow_card_value.get()
                
                info_obj["Red Card"] -= 2 * red_card_value.get()
                
                if fan_choice.get() == "Few Fan":
                    info_obj["Fans"] += 1
                elif fan_choice.get() == "Huge Fan":
                    info_obj[">3fans"] += 2
                
                if register_choice.get() == "Not Registered":
                    info_obj["Ko ĐK"] -= 1
                
                if travel_choice.get() == "Go":
                    info_obj["Đi"] += 1
                    info_obj["Liên tiếp"] += 1
                    if register_choice.get() == "Registered":
                        pass
                    elif register_choice.get() == "Not Registered":
                        info_obj["oĐK Đi"] -= 1 
                elif travel_choice.get() == "Late Go":
                    info_obj["Đi"] += 1
                    info_obj["Liên tiếp"] += 1
                    info_obj["Đi Muộn"] -= 1
                    if register_choice.get() == "Registered":
                        pass
                    elif register_choice.get() == "Not Registered":
                        info_obj["oĐK Đi"] -= 1 
                elif travel_choice.get() == "Not Go":
                    info_obj["Ko Đi"] -= 1
                    info_obj["Liên tiếp"] = 0
                    if register_choice.get() == "Registered":
                        info_obj["ĐK oĐi"] -= 1
                elif travel_choice.get() == "DK Not Go":
                    info_obj["Ko Đi"] -= 1
                    info_obj["Liên tiếp"] = 0
                
                if long_ho != 2:
                    info_obj["Vòng"] = (info_obj["Điểm cấp"]    
                                        + info_obj["Đi"]
                                        + info_obj["Ko Đi"]
                                        + info_obj["Đi Muộn"]
                                        + info_obj["Ko ĐK"]
                                        + info_obj["ĐK oĐi"]
                                        + info_obj["oĐK Đi"]
                                        + info_obj["Chấn Thg"]
                                        + info_obj["Bạn CTg"]
                                        + info_obj["Áo Ko Đúng"]
                                        + info_obj["Phí muộn"]
                                        + info_obj["Yellow Card"]
                                        + info_obj["Red Card"]
                                        + info_obj["Giám hộ"]
                                        + info_obj["Fans"]
                                        + info_obj[">3fans"]
                                        + info_obj["Chuối"]
                                        + info_obj["Hattrick"]
                                        )       
                
                if long_ho != 2:
                    info_obj["KVH Index"] = info_obj["Vòng"] + info_obj["Thưởng thắng"]
                
                if win_award_checkbox_value.get() and coach_checkbox_value.get():
                    info_obj["KVH Index"] += 1    
                    
                return info_obj
        return {}

    # Hàm cập nhật thông tin trong info_frame
    def update_info(info):
        # Lưu trạng thái cuộn hiện tại
        scroll_position = info_text.yview()

        info_text.config(state='normal')  # Cho phép chỉnh sửa trạng thái
        info_text.delete(1.0, tk.END)  # Xóa nội dung cũ trước khi thêm mới
        info_text.insert(tk.END, info)
        info_text.config(state='disabled')  # Khóa lại trạng thái

        # Thiết lập lại trạng thái cuộn
        info_text.yview_moveto(scroll_position[0])
        
            
    def update_info_frame():
        selected_name = combo_box.get()
        info_obj = get_info(selected_name)
        if info_obj != {}:
            info = get_info_from_obj(info_obj)
            update_info(info)

    # Tạo radio button Đi
    radio_go = ttk.Radiobutton(radio_frame, text="Đi", variable=travel_choice, value="Go", command=update_info_frame)
    radio_go.pack(anchor="w", padx=5)

    # Tạo radio button Đi muộn
    radio_late_go = ttk.Radiobutton(radio_frame, text="Đi muộn", variable=travel_choice, value="Late Go", command=update_info_frame)
    radio_late_go.pack(anchor="w", padx=5)

    # Tạo radio button Không đi
    radio_not_go = ttk.Radiobutton(radio_frame, text="Không đi", variable=travel_choice, value="Not Go", command=update_info_frame)
    radio_not_go.pack(anchor="w", padx=5)
    
    # Tạo radio button ĐK là Không đi
    radio_DK_not_go = ttk.Radiobutton(radio_frame, text="ĐK là Không đi", variable=travel_choice, value="DK Not Go", command=update_info_frame)
    radio_DK_not_go.pack(anchor="w", padx=5)

    # Tạo khoảng trống giữa hai tùy chọn
    ttk.Label(radio_frame, text="").pack()

    # Tạo radio button Đăng ký
    radio_register = ttk.Radiobutton(radio_frame, text="Đăng ký", variable=register_choice, value="Registered", command=update_info_frame)
    radio_register.pack(anchor="w", padx=5)

    # Tạo radio button Không đăng ký
    radio_not_register = ttk.Radiobutton(radio_frame, text="Không đăng ký", variable=register_choice, value="Not Registered", command=update_info_frame)
    radio_not_register.pack(anchor="w", padx=5)
    
     # Tạo khoảng trống giữa hai tùy chọn
    ttk.Label(radio_frame, text="").pack()
    
    # Tạo radio button không có fan
    radio_no_fan = ttk.Radiobutton(radio_frame, text="Không có fan", variable=fan_choice, value="No Fan", command=update_info_frame)
    radio_no_fan.pack(anchor="w", padx=5)

    # Tạo radio button có 1 đến 3 fan
    radio_few_fan = ttk.Radiobutton(radio_frame, text="Có 1 đến 3 fans", variable=fan_choice, value="Few Fan", command=update_info_frame)
    radio_few_fan.pack(anchor="w", padx=5)

    # Tạo radio button có > 3 fan
    radio_huge_fan = ttk.Radiobutton(radio_frame, text="Có > 3 fans", variable=fan_choice, value="Huge Fan", command=update_info_frame)
    radio_huge_fan.pack(anchor="w", padx=5)
    
    # Tạo khoảng trống giữa hai tùy chọn
    ttk.Label(radio_frame, text="").pack()
    
    # Thêm checkbox cho trường chấn thương
    injury_checkbox = ttk.Checkbutton(radio_frame, text="Chấn thương", variable=injury_checkbox_value, command=update_info_frame)
    injury_checkbox.pack(anchor="w", padx=5)
    
    # Thêm checkbox cho trường bạn chấn thương
    self_injury = ttk.Checkbutton(radio_frame, text="Bạn Chấn thương", variable=self_injury_value, command=update_info_frame)
    self_injury.pack(anchor="w", padx=5)
    
    # Thêm checkbox cho trường áo không đúng
    wrong_clothes_checkbox = ttk.Checkbutton(radio_frame, text="Áo không đúng", variable=wrong_clothes_checkbox_value, command=update_info_frame)
    wrong_clothes_checkbox.pack(anchor="w", padx=5)
    
    # Thêm checkbox cho trường phí muộn
    late_fee_checkbox = ttk.Checkbutton(radio_frame, text="Phí muộn", variable=late_fee_checkbox_value, command=update_info_frame)
    late_fee_checkbox.pack(anchor="w", padx=5)
    
    # Thêm checkbox cho trường giám hộ
    supervisor_checkbox = ttk.Checkbutton(radio_frame, text="Giám hộ", variable=supervisor_checkbox_value, command=update_info_frame)
    supervisor_checkbox.pack(anchor="w", padx=5)
    
    # Thêm checkbox cho trường chuối
    banana_checkbox = ttk.Checkbutton(radio_frame, text="Chuối", variable=banana_checkbox_value, command=update_info_frame)
    banana_checkbox.pack(anchor="w", padx=5)
    
    # Thêm checkbox cho trường hattrick
    hattrick_checkbox = ttk.Checkbutton(radio_frame, text="Hattrick", variable=hattrick_checkbox_value, command=update_info_frame)
    hattrick_checkbox.pack(anchor="w", padx=5)
    
    # Thêm checkbox cho trường Thưởng thắng
    win_award_checkbox = ttk.Checkbutton(radio_frame, text="Thưởng thắng", variable=win_award_checkbox_value, command=update_info_frame)
    win_award_checkbox.pack(anchor="w", padx=5)
    
    # Thêm checkbox cho trường HLV
    coach_checkbox = ttk.Checkbutton(radio_frame, text="HLV", variable=coach_checkbox_value, command=update_info_frame)
    coach_checkbox.pack(anchor="w", padx=5)
    
    def increase_goal():
        goal_value.set(goal_value.get() + 1)
        update_info_frame()

    def decrease_goal():
        goal_value.set(max(0, goal_value.get() - 1))
        update_info_frame()
        
    def increase_yellow_card():
        yellow_card_value.set(yellow_card_value.get() + 1)
        update_info_frame()

    def decrease_yellow_card():
        yellow_card_value.set(max(0, yellow_card_value.get() - 1))
        update_info_frame()
        
    def increase_red_card():
        red_card_value.set(red_card_value.get() + 1)
        update_info_frame()

    def decrease_red_card():
        red_card_value.set(max(0, red_card_value.get() - 1))
        update_info_frame()
    
    # Tạo frame chứa trường bàn thắng
    goal_frame = ttk.Frame(radio_frame)
    goal_frame.pack(anchor="w", pady=(10, 0))

    # Thêm trường bàn thắng
    ttk.Label(goal_frame, text="Bàn thắng:", anchor="w").pack(side="left")
    ttk.Button(goal_frame, text="-", command=decrease_goal, width=3).pack(side="left", padx=(5, 0))
    ttk.Label(goal_frame, textvariable=goal_value).pack(side="left", padx=(5, 5))
    ttk.Button(goal_frame, text="+", command=increase_goal, width=3).pack(side="left")
    
    # Tạo frame chứa trường thẻ vàng
    yellow_card_frame = ttk.Frame(radio_frame)
    yellow_card_frame.pack(anchor="w", pady=(10, 0))

    # Thêm trường thẻ vàng
    ttk.Label(yellow_card_frame, text="Thẻ vàng:", anchor="w").pack(side="left")
    ttk.Button(yellow_card_frame, text="-", command=decrease_yellow_card, width=3).pack(side="left", padx=(5, 0))
    ttk.Label(yellow_card_frame, textvariable=yellow_card_value).pack(side="left", padx=(5, 5))
    ttk.Button(yellow_card_frame, text="+", command=increase_yellow_card, width=3).pack(side="left")

    # Tạo frame chứa trường thẻ đỏ
    red_card_frame = ttk.Frame(radio_frame)
    red_card_frame.pack(anchor="w", pady=(10, 0))

    # Thêm trường thẻ đỏ
    ttk.Label(red_card_frame, text="Thẻ đỏ:", anchor="w").pack(side="left")
    ttk.Button(red_card_frame, text="-", command=decrease_red_card, width=3).pack(side="left", padx=(5, 0))
    ttk.Label(red_card_frame, textvariable=red_card_value).pack(side="left", padx=(5, 5))
    ttk.Button(red_card_frame, text="+", command=increase_red_card, width=3).pack(side="left")

    save_frame = ttk.Frame(radio_frame)
    save_frame.pack(anchor="w", pady=(10, 0))

    def update_global_info(selected_name, info_obj):
        for i in range(len(data_dict)):
            name = data_dict[i]["Họ và tên"]
            if name == selected_name:
                data_dict[i] = info_obj
                return

    def obj_to_tuple(obj):
        # Sắp xếp các giá trị theo thứ tự key và chuyển đổi thành tuple
        return tuple(obj[key] for key in obj.keys())
    
    def save_info():
        selected_name = combo_box.get()
        info_obj = get_info(selected_name)
        if info_obj != {}:
            confirm = messagebox.askyesno("Xác nhận", "Bạn có chắc muốn lưu thông tin này?")      
            if confirm:
                update_global_info(selected_name, info_obj) 
                for row in treeview.get_children():
                    name = treeview.item(row, "values")[0]
                    if name == selected_name:
                        treeview.item(row, values=obj_to_tuple(info_obj))
                        # Tô màu hàng vừa lưu
                        treeview.tag_configure("saved_row", background="blue")
                        treeview.item(row, tags=("saved_row",))
                # Xóa tên người vừa lưu khỏi Combobox
                index_to_delete = combo_box["values"].index(selected_name)
                combo_box["values"] = combo_box["values"][:index_to_delete] + combo_box["values"][index_to_delete+1:]
        else:
            messagebox.showwarning("Cảnh báo", "Vui lòng chọn tên cầu thủ trước!")
            return


    # Tạo một nút để lưu dữ liệu và đặt vào frame lưu nút
    save_button = ttk.Button(save_frame, text="Lưu", command=save_info)
    save_button.pack(padx=70, pady=10)
    
# Tạo một cửa sổ
root = tk.Tk()
root.title("Quản lý điểm KVH Long ho")

# Đặt kích thước và vị trí của cửa sổ (kích thước 800x600 và đặt ở giữa màn hình)
window_width = 800
window_height = 600
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x_coordinate = (screen_width - window_width) // 2
y_coordinate = (screen_height - window_height) // 2
root.geometry(f"{window_width}x{window_height}+{x_coordinate}+{y_coordinate}")

# Tạo một khung chứa Treeview
frame = ttk.Frame(root)
frame.pack(fill="both", expand=True)

# Tạo một Treeview
treeview = ttk.Treeview(frame)

# Tạo thanh cuộn dọc
vertical_scrollbar = ttk.Scrollbar(frame, orient="vertical", command=treeview.yview)
treeview.configure(yscrollcommand=vertical_scrollbar.set)
vertical_scrollbar.pack(side="right", fill="y")

# Tạo thanh cuộn ngang
horizontal_scrollbar = ttk.Scrollbar(frame, orient="horizontal", command=treeview.xview)
treeview.configure(xscrollcommand=horizontal_scrollbar.set)
horizontal_scrollbar.pack(side="bottom", fill="x")

# Hiển thị Treeview
treeview.pack(expand=True, fill="both")

# Tạo một frame con để chứa các nút
button_frame = ttk.Frame(root)
button_frame.pack()

# Tạo một button để import file Excel và đặt vào frame con
import_button = ttk.Button(button_frame, text="Nhập file Excel", command=import_excel)
import_button.grid(row=0, column=0, padx=10, pady=10)

# Tạo một button để option và đặt vào frame con
option_button = ttk.Button(button_frame, text="Tùy chọn", command=option)
option_button.grid(row=0, column=1, padx=10, pady=10)

# Tạo một button để import file Excel và đặt vào frame con
export_button = ttk.Button(button_frame, text="Xuất file Excel", command=export_excel)
export_button.grid(row=0, column=2, padx=10, pady=10)

# Tạo một button để mở cửa sổ thông kê
statistics_button = ttk.Button(button_frame, text="Thông kê", command=open_statistics_window)
statistics_button.grid(row=0, column=3, padx=10, pady=10)

# Tạo một button để sắp xếp
sort_button = ttk.Button(button_frame, text="Sắp xếp", command=sort_data_by_kvh_index)
sort_button.grid(row=0, column=4, padx=10, pady=10)

# Tạo style cho frame chứa thông tin
style = ttk.Style()
style.configure("Info.TFrame", background="white")

# Tạo các tag với màu nền và màu chữ tương ứng
treeview.tag_configure("purple", background="purple")
treeview.tag_configure("red", background="red")
treeview.tag_configure("green", background="green")
treeview.tag_configure("pink", background="pink")
treeview.tag_configure("lightblue", background="lightblue")
treeview.tag_configure("yellow", background="yellow")
treeview.tag_configure("white_text", foreground="white")


# Chạy vòng lặp chính
root.mainloop()